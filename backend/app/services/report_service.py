"""
Report Generation Service
Generates PDF and Excel reports with charts and statistics
Aligned with dashboard analytics, filters, and system settings
"""
import os
import io
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract
from collections import Counter, defaultdict
from math import sqrt

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display

# Charts
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# Excel Generation
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, LineChart, Reference

from app.models.feedback import Feedback
from app.models.report import Report, ReportStatus


@dataclass
class AnalyticsSettings:
    """Analytics settings that control report calculations"""
    nps_target: int = 50
    csat_threshold: int = 80
    min_reviews_per_route: int = 10
    
    @classmethod
    def from_dict(cls, data: Optional[Dict[str, Any]] = None) -> 'AnalyticsSettings':
        """Create settings from a dictionary"""
        if not data:
            return cls()
        return cls(
            nps_target=data.get('nps_target', 50),
            csat_threshold=data.get('csat_threshold', 80),
            min_reviews_per_route=data.get('min_reviews_per_route', 10)
        )


# Register Arabic font
def register_arabic_font():
    """Register an Arabic-compatible font for PDF generation"""
    try:
        # Try common system fonts that support Arabic
        font_paths = [
            # Windows fonts
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/tahoma.ttf",
            "C:/Windows/Fonts/segoeui.ttf",
            # Linux fonts
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Arabic', font_path))
                return 'Arabic'
        
        # Fallback to Helvetica if no Arabic font found
        return 'Helvetica'
    except Exception:
        return 'Helvetica'


def reshape_arabic_text(text: str) -> str:
    """Reshape Arabic text for proper display in PDF"""
    if not text:
        return text
    
    try:
        # Check if text contains Arabic characters
        has_arabic = any('\u0600' <= c <= '\u06FF' or '\u0750' <= c <= '\u077F' for c in text)
        
        if has_arabic:
            reshaped = arabic_reshaper.reshape(text)
            return get_display(reshaped)
        return text
    except Exception:
        return text


class ReportService:
    """Service for generating comprehensive feedback reports aligned with dashboard analytics"""
    
    def __init__(self, db: Session, analytics_settings: Optional[AnalyticsSettings] = None):
        self.db = db
        self.reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'reports')
        os.makedirs(self.reports_dir, exist_ok=True)
        self.arabic_font = register_arabic_font()
        self.settings = analytics_settings or AnalyticsSettings()
    
    def validate_data_availability(self, feedbacks: List[Feedback], stats: Dict[str, Any]) -> Dict[str, Any]:
        """
        Validate that sufficient data is available for report generation.
        Returns validation result with warnings for missing data.
        """
        validation = {
            'is_valid': True,
            'warnings': [],
            'errors': [],
            'available_sections': []
        }
        
        total = stats.get('total', 0)
        
        # Critical: Must have at least some feedback
        if total == 0:
            validation['is_valid'] = False
            validation['errors'].append('No feedback data matches the selected filters')
            return validation
        
        # Check for sentiment data
        if stats.get('positive', 0) + stats.get('negative', 0) + stats.get('neutral', 0) == 0:
            validation['warnings'].append('No sentiment data available - NPS and CSAT cannot be calculated')
        else:
            validation['available_sections'].extend(['sentiment_overview', 'nps', 'csat'])
        
        # Check for trend data (need at least 2 different dates)
        dates = set(f.feedback_date.strftime('%Y-%m-%d') for f in feedbacks if f.feedback_date)
        if len(dates) < 2:
            validation['warnings'].append('Insufficient date range for trend analysis (need at least 2 dates)')
        else:
            validation['available_sections'].append('trend_analysis')
        
        # Check for route data
        routes_with_min_reviews = sum(1 for _ in self._get_route_counts(feedbacks) 
                                       if _ >= self.settings.min_reviews_per_route)
        if routes_with_min_reviews == 0:
            validation['warnings'].append(
                f'No routes meet minimum review requirement ({self.settings.min_reviews_per_route} reviews)'
            )
        else:
            validation['available_sections'].append('route_performance')
        
        # Check for language distribution
        if stats.get('arabic_count', 0) + stats.get('english_count', 0) > 0:
            validation['available_sections'].append('language_distribution')
        
        return validation
    
    def _get_route_counts(self, feedbacks: List[Feedback]) -> List[int]:
        """Helper to get counts of feedback per route"""
        route_counts = defaultdict(int)
        for f in feedbacks:
            if f.flight_number:
                route_counts[f.flight_number] += 1
        return list(route_counts.values())
    
    def get_filtered_feedback(
        self,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        sentiments: Optional[List[str]] = None,
        languages: Optional[List[str]] = None
    ) -> Tuple[List[Feedback], Dict[str, Any]]:
        """
        Get feedback data with applied filters and calculate statistics
        """
        query = self.db.query(Feedback)
        
        # Apply date filters
        if date_from:
            query = query.filter(Feedback.feedback_date >= date_from)
        if date_to:
            # Include the entire end date
            end_of_day = date_to.replace(hour=23, minute=59, second=59)
            query = query.filter(Feedback.feedback_date <= end_of_day)
        
        # Apply sentiment filters
        if sentiments and len(sentiments) < 3:
            query = query.filter(Feedback.sentiment.in_(sentiments))
        
        # Apply language filters
        if languages:
            lang_conditions = []
            if 'AR' in languages or 'arabic' in languages:
                lang_conditions.append(Feedback.language == 'AR')
            if 'EN' in languages or 'english' in languages:
                lang_conditions.append(Feedback.language == 'EN')
            if lang_conditions:
                query = query.filter(and_(*lang_conditions) if len(lang_conditions) == 1 else Feedback.language.in_(['AR', 'EN']))
        
        feedbacks = query.order_by(Feedback.feedback_date.desc()).all()
        
        # Calculate statistics
        total = len(feedbacks)
        positive = sum(1 for f in feedbacks if f.sentiment == 'positive')
        negative = sum(1 for f in feedbacks if f.sentiment == 'negative')
        neutral = sum(1 for f in feedbacks if f.sentiment == 'neutral')
        
        ar_count = sum(1 for f in feedbacks if f.language == 'AR')
        en_count = sum(1 for f in feedbacks if f.language == 'EN')
        mixed_count = sum(1 for f in feedbacks if f.language == 'Mixed')
        
        stats = {
            'total': total,
            'positive': positive,
            'negative': negative,
            'neutral': neutral,
            'positive_pct': round(positive / total * 100, 1) if total > 0 else 0,
            'negative_pct': round(negative / total * 100, 1) if total > 0 else 0,
            'neutral_pct': round(neutral / total * 100, 1) if total > 0 else 0,
            'arabic_count': ar_count,
            'english_count': en_count,
            'mixed_count': mixed_count,
            'avg_confidence': round(sum(f.sentiment_confidence or 0 for f in feedbacks) / total, 2) if total > 0 else 0,
        }
        
        return feedbacks, stats
    
    def get_trend_data(self, feedbacks: List[Feedback]) -> Dict[str, List]:
        """
        Aggregate feedback by date for trend analysis
        """
        daily_data = defaultdict(lambda: {'positive': 0, 'negative': 0, 'neutral': 0, 'total': 0})
        
        for f in feedbacks:
            date_key = f.feedback_date.strftime('%Y-%m-%d') if f.feedback_date else 'Unknown'
            daily_data[date_key][f.sentiment or 'neutral'] += 1
            daily_data[date_key]['total'] += 1
        
        # Sort by date
        sorted_dates = sorted(daily_data.keys())
        
        return {
            'dates': sorted_dates,
            'positive': [daily_data[d]['positive'] for d in sorted_dates],
            'negative': [daily_data[d]['negative'] for d in sorted_dates],
            'neutral': [daily_data[d]['neutral'] for d in sorted_dates],
            'total': [daily_data[d]['total'] for d in sorted_dates],
        }
    
    def get_top_negative_feedback(self, feedbacks: List[Feedback], limit: int = 10) -> List[Dict]:
        """
        Get top negative feedback samples sorted by confidence
        """
        negative_feedbacks = [
            f for f in feedbacks 
            if f.sentiment == 'negative'
        ]
        
        # Sort by confidence (highest first)
        negative_feedbacks.sort(key=lambda x: x.sentiment_confidence or 0, reverse=True)
        
        return [
            {
                'id': f.id,
                'text': f.text[:200] + '...' if len(f.text) > 200 else f.text,
                'date': f.feedback_date.strftime('%Y-%m-%d') if f.feedback_date else 'N/A',
                'confidence': f.sentiment_confidence or 0,
                'flight': f.flight_number or 'N/A',
                'customer': f.customer_name or 'Anonymous'
            }
            for f in negative_feedbacks[:limit]
        ]
    
    def calculate_nps(self, stats: Dict[str, Any]) -> Dict[str, Any]:
        """
        Calculate Net Promoter Score (NPS) based on sentiment distribution.
        
        NPS Mapping:
        - Positive sentiment → Promoters (9-10)
        - Neutral sentiment → Passives (7-8)
        - Negative sentiment → Detractors (0-6)
        
        NPS = % Promoters - % Detractors (range: -100 to +100)
        """
        total = stats['total']
        if total == 0:
            return {
                'nps_score': 0,
                'grade': 'N/A',
                'promoter_pct': 0,
                'passive_pct': 0,
                'detractor_pct': 0
            }
        
        promoter_pct = round((stats['positive'] / total) * 100, 1)
        passive_pct = round((stats['neutral'] / total) * 100, 1)
        detractor_pct = round((stats['negative'] / total) * 100, 1)
        
        nps_score = round(promoter_pct - detractor_pct)
        
        # NPS Grade
        if nps_score >= 70:
            grade = 'Excellent'
        elif nps_score >= 50:
            grade = 'Great'
        elif nps_score >= 30:
            grade = 'Good'
        elif nps_score >= 0:
            grade = 'Needs Improvement'
        else:
            grade = 'Critical'
        
        # Compare with target
        target = self.settings.nps_target
        vs_target = nps_score - target
        meets_target = nps_score >= target
        
        return {
            'nps_score': nps_score,
            'grade': grade,
            'promoter_pct': promoter_pct,
            'passive_pct': passive_pct,
            'detractor_pct': detractor_pct,
            'target': target,
            'vs_target': vs_target,
            'meets_target': meets_target
        }
    
    def calculate_csat(self, stats: Dict[str, Any]) -> Dict[str, Any]:
        """
        Calculate Customer Satisfaction Score (CSAT).
        
        CSAT = Positive / Total * 100 (matching dashboard calculation)
        Using configurable threshold from analytics settings.
        """
        total = stats['total']
        if total == 0:
            return {
                'csat_score': 0,
                'threshold': self.settings.csat_threshold,
                'meets_threshold': False,
                'satisfied_count': 0,
                'unsatisfied_count': 0,
                'status': 'N/A'
            }
        
        # Satisfied = Positive only (matching dashboard calculation)
        satisfied_count = stats['positive']
        unsatisfied_count = stats['negative']
        csat_score = round((satisfied_count / total) * 100, 1)
        
        threshold = self.settings.csat_threshold
        meets_threshold = csat_score >= threshold
        
        # Status based on CSAT score
        if csat_score >= 90:
            status = 'Excellent'
        elif csat_score >= 80:
            status = 'Good'
        elif csat_score >= 70:
            status = 'Fair'
        elif csat_score >= 60:
            status = 'Needs Improvement'
        else:
            status = 'Critical'
        
        return {
            'csat_score': csat_score,
            'threshold': threshold,
            'meets_threshold': meets_threshold,
            'satisfied_count': satisfied_count,
            'unsatisfied_count': unsatisfied_count,
            'status': status
        }
    
    def get_monthly_nps(self, feedbacks: List[Feedback]) -> List[Dict[str, Any]]:
        """
        Calculate NPS grouped by month for trend analysis.
        Returns list of monthly NPS data sorted by date.
        """
        monthly_data = defaultdict(lambda: {'positive': 0, 'negative': 0, 'neutral': 0, 'total': 0})
        
        for f in feedbacks:
            if f.feedback_date:
                month_key = f.feedback_date.strftime('%Y-%m')
                sentiment = f.sentiment or 'neutral'
                monthly_data[month_key][sentiment] += 1
                monthly_data[month_key]['total'] += 1
        
        results = []
        for month, data in sorted(monthly_data.items()):
            total = data['total']
            if total > 0:
                promoter_pct = (data['positive'] / total) * 100
                detractor_pct = (data['negative'] / total) * 100
                nps_score = round(promoter_pct - detractor_pct)
                
                results.append({
                    'month': month,
                    'month_display': datetime.strptime(month, '%Y-%m').strftime('%b %Y'),
                    'nps_score': nps_score,
                    'total': total,
                    'positive': data['positive'],
                    'negative': data['negative'],
                    'neutral': data['neutral'],
                    'meets_target': nps_score >= self.settings.nps_target
                })
        
        return results
    
    def get_complaint_categories(self, feedbacks: List[Feedback], limit: int = 10) -> List[Dict[str, Any]]:
        """
        Extract and rank top complaint categories from negative feedback.
        Analyzes feedback text using keyword detection.
        """
        keyword_counts = Counter()
        
        # Common complaint keywords for detection
        complaint_keywords = {
            'delay': ['delay', 'late', 'wait', 'waiting', 'delayed', 'تأخير', 'متأخر'],
            'baggage': ['baggage', 'luggage', 'bag', 'bags', 'lost', 'حقيبة', 'أمتعة'],
            'service': ['service', 'staff', 'crew', 'rude', 'unfriendly', 'خدمة', 'طاقم'],
            'food': ['food', 'meal', 'catering', 'hungry', 'طعام', 'وجبة'],
            'seating': ['seat', 'seating', 'cramped', 'uncomfortable', 'مقعد', 'جلوس'],
            'booking': ['booking', 'reservation', 'ticket', 'cancel', 'حجز', 'تذكرة'],
            'refund': ['refund', 'money', 'compensation', 'payment', 'استرداد', 'مال'],
            'cleanliness': ['clean', 'dirty', 'hygiene', 'sanitary', 'نظافة', 'قذر'],
            'entertainment': ['entertainment', 'screen', 'wifi', 'internet', 'ترفيه', 'شاشة'],
            'check-in': ['check-in', 'checkin', 'counter', 'airport', 'تسجيل', 'مطار']
        }
        
        negative_feedbacks = [f for f in feedbacks if f.sentiment == 'negative']
        
        for f in negative_feedbacks:
            # Analyze text for keywords to detect complaint categories
            if f.text:
                text_lower = f.text.lower()
                for category, keywords in complaint_keywords.items():
                    if any(kw in text_lower for kw in keywords):
                        keyword_counts[category] += 1
        
        total_negative = len(negative_feedbacks) or 1
        
        results = []
        for category, count in keyword_counts.most_common(limit):
            results.append({
                'category': category.title(),
                'count': count,
                'percentage': round((count / total_negative) * 100, 1)
            })
        
        return results
    
    def get_top_routes(self, feedbacks: List[Feedback], limit: int = 5) -> List[Dict]:
        """
        Get top routes with Wilson Score ranking for statistical confidence.
        Uses min_reviews_per_route from analytics settings.
        """
        route_data = defaultdict(lambda: {'positive': 0, 'negative': 0, 'neutral': 0, 'total': 0})
        
        for f in feedbacks:
            # Use flight_number as route identifier
            route = f.flight_number or 'Unknown'
            if route == 'Unknown':
                continue  # Skip feedbacks without flight numbers
            route_data[route][f.sentiment or 'neutral'] += 1
            route_data[route]['total'] += 1
        
        # Use configurable minimum sample size from settings
        min_reviews = self.settings.min_reviews_per_route
        
        # Calculate Wilson Score for each route
        routes_with_scores = []
        for route, data in route_data.items():
            if data['total'] >= min_reviews:
                n = data['total']
                p = data['positive'] / n
                z = 1.96  # 95% confidence
                
                # Wilson Score lower bound
                wilson_score = (p + z*z/(2*n) - z * sqrt((p*(1-p) + z*z/(4*n))/n)) / (1 + z*z/n)
                
                # Calculate average sentiment score (-1 to +1)
                sentiment_score = ((data['positive'] - data['negative']) / n)
                
                # Determine confidence level based on sample size
                if n >= 100:
                    confidence_level = 'Very High'
                elif n >= 50:
                    confidence_level = 'High'
                elif n >= 20:
                    confidence_level = 'Medium'
                else:
                    confidence_level = 'Low'
                
                routes_with_scores.append({
                    'route': route,
                    'positive': data['positive'],
                    'negative': data['negative'],
                    'neutral': data['neutral'],
                    'total': data['total'],
                    'positive_rate': round(p * 100, 1),
                    'wilson_score': round(wilson_score * 100, 1),
                    'sentiment_score': round(sentiment_score * 100, 1),
                    'confidence_level': confidence_level
                })
        
        # Sort by Wilson Score
        routes_with_scores.sort(key=lambda x: x['wilson_score'], reverse=True)
        return routes_with_scores[:limit]
    
    def create_sentiment_chart(self, stats: Dict[str, Any], width: int = 400, height: int = 300) -> bytes:
        """
        Create a pie chart for sentiment distribution
        """
        fig, ax = plt.subplots(figsize=(width/100, height/100), dpi=100)
        
        sizes = [stats['positive'], stats['negative'], stats['neutral']]
        labels = ['Positive', 'Negative', 'Neutral']
        colors_list = ['#22C55E', '#EF4444', '#F59E0B']
        explode = (0.02, 0.02, 0.02)
        
        # Only show non-zero values
        non_zero = [(s, l, c) for s, l, c in zip(sizes, labels, colors_list) if s > 0]
        if non_zero:
            sizes, labels, colors_list = zip(*non_zero)
        
        ax.pie(sizes, explode=explode[:len(sizes)], labels=labels, colors=colors_list,
               autopct='%1.1f%%', shadow=False, startangle=90)
        ax.axis('equal')
        ax.set_title('Sentiment Distribution', fontsize=14, fontweight='bold')
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight', dpi=100)
        plt.close()
        buf.seek(0)
        return buf.getvalue()
    
    def create_trend_chart(self, trend_data: Dict[str, List], width: int = 600, height: int = 300) -> bytes:
        """
        Create a line chart for sentiment trends over time
        """
        fig, ax = plt.subplots(figsize=(width/100, height/100), dpi=100)
        
        dates = trend_data['dates']
        
        if len(dates) > 30:
            # Show only every nth date if too many
            step = len(dates) // 15
            dates_display = dates[::step]
        else:
            dates_display = dates
        
        ax.plot(range(len(dates)), trend_data['positive'], 'g-', label='Positive', linewidth=2)
        ax.plot(range(len(dates)), trend_data['negative'], 'r-', label='Negative', linewidth=2)
        ax.plot(range(len(dates)), trend_data['neutral'], 'orange', label='Neutral', linewidth=1.5, alpha=0.7)
        
        ax.set_xlabel('Date')
        ax.set_ylabel('Count')
        ax.set_title('Sentiment Trends Over Time', fontsize=14, fontweight='bold')
        ax.legend(loc='upper left')
        ax.grid(True, alpha=0.3)
        
        # Set x-axis labels
        if len(dates) > 0:
            tick_positions = list(range(0, len(dates), max(1, len(dates) // 10)))
            ax.set_xticks(tick_positions)
            ax.set_xticklabels([dates[i][-5:] for i in tick_positions], rotation=45, ha='right')
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight', dpi=100)
        plt.close()
        buf.seek(0)
        return buf.getvalue()
    
    def create_language_chart(self, stats: Dict[str, Any], width: int = 300, height: int = 200) -> bytes:
        """
        Create a bar chart for language distribution
        """
        fig, ax = plt.subplots(figsize=(width/100, height/100), dpi=100)
        
        languages = ['Arabic', 'English', 'Mixed']
        counts = [stats['arabic_count'], stats['english_count'], stats.get('mixed_count', 0)]
        colors_list = ['#003366', '#C5A572', '#14B8A6']
        
        # Only show languages with non-zero counts
        non_zero = [(l, c, col) for l, c, col in zip(languages, counts, colors_list) if c > 0]
        if non_zero:
            languages, counts, colors_list = zip(*non_zero)
        
        bars = ax.bar(languages, counts, color=colors_list)
        ax.set_ylabel('Count')
        ax.set_title('Language Distribution', fontsize=12, fontweight='bold')
        
        # Add value labels on bars
        for bar, count in zip(bars, counts):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, 
                   str(count), ha='center', va='bottom')
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight', dpi=100)
        plt.close()
        buf.seek(0)
        return buf.getvalue()
    
    def create_monthly_nps_chart(self, monthly_data: List[Dict], width: int = 600, height: int = 300) -> bytes:
        """
        Create a line chart showing monthly NPS trend with target line.
        """
        fig, ax = plt.subplots(figsize=(width/100, height/100), dpi=100)
        
        months = [m['month_display'] for m in monthly_data]
        nps_scores = [m['nps_score'] for m in monthly_data]
        target = self.settings.nps_target
        
        # NPS Score line
        ax.plot(range(len(months)), nps_scores, 'b-', label='NPS Score', linewidth=2, marker='o', markersize=6)
        
        # Target line
        ax.axhline(y=target, color='#22C55E', linestyle='--', linewidth=1.5, label=f'Target ({target})')
        
        # Color points based on whether they meet target
        for i, (score, meets) in enumerate(zip(nps_scores, [m['meets_target'] for m in monthly_data])):
            color = '#22C55E' if meets else '#EF4444'
            ax.scatter(i, score, color=color, s=80, zorder=5)
        
        ax.set_xlabel('Month')
        ax.set_ylabel('NPS Score')
        ax.set_title('Monthly NPS Trend', fontsize=14, fontweight='bold')
        ax.legend(loc='upper left')
        ax.grid(True, alpha=0.3)
        
        # Set y-axis range
        min_score = min(nps_scores) if nps_scores else 0
        max_score = max(nps_scores) if nps_scores else 100
        ax.set_ylim(min(min_score - 10, -10), max(max_score + 10, target + 20))
        
        # Set x-axis labels
        ax.set_xticks(range(len(months)))
        ax.set_xticklabels(months, rotation=45, ha='right')
        
        # Add score labels
        for i, score in enumerate(nps_scores):
            ax.annotate(str(score), (i, score), textcoords="offset points", 
                       xytext=(0, 10), ha='center', fontsize=9)
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight', dpi=100)
        plt.close()
        buf.seek(0)
        return buf.getvalue()
    
    def generate_pdf_report(
        self,
        title: str,
        feedbacks: List[Feedback],
        stats: Dict[str, Any],
        date_from: Optional[datetime],
        date_to: Optional[datetime],
        sections: Dict[str, bool],
        include_logo: bool = True,
        orientation: str = 'portrait'
    ) -> Tuple[str, int]:
        """
        Generate a PDF report with charts and statistics
        Returns: (file_path, file_size)
        """
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.reports_dir, filename)
        
        # Set page size
        if orientation == 'landscape':
            pagesize = landscape(letter)
        else:
            pagesize = letter
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=pagesize,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#003366'),
            spaceAfter=12,
            alignment=1  # Center
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#003366'),
            spaceBefore=20,
            spaceAfter=10
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=6
        )
        
        # Title
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        # Date range
        date_range_text = ""
        if date_from and date_to:
            date_range_text = f"Period: {date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        elif date_from:
            date_range_text = f"From: {date_from.strftime('%Y-%m-%d')}"
        elif date_to:
            date_range_text = f"Until: {date_to.strftime('%Y-%m-%d')}"
        
        if date_range_text:
            story.append(Paragraph(date_range_text, normal_style))
        
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", normal_style))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        if sections.get('executiveSummary', True):
            story.append(Paragraph("Executive Summary", heading_style))
            
            summary_data = [
                ['Metric', 'Value', 'Percentage'],
                ['Total Feedback', str(stats['total']), '100%'],
                ['Positive Feedback', str(stats['positive']), f"{stats['positive_pct']}%"],
                ['Negative Feedback', str(stats['negative']), f"{stats['negative_pct']}%"],
                ['Neutral Feedback', str(stats['neutral']), f"{stats['neutral_pct']}%"],
                ['Average Confidence', f"{stats['avg_confidence']:.2f}", '-'],
            ]
            
            table = Table(summary_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F7FA')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
            ]))
            story.append(table)
            story.append(Spacer(1, 20))
        
        # NPS Score Section
        if sections.get('npsScore', True):
            story.append(Paragraph("Net Promoter Score (NPS)", heading_style))
            
            nps_data = self.calculate_nps(stats)
            
            # NPS score color based on grade
            nps_color = colors.HexColor('#22C55E')  # Green
            if nps_data['nps_score'] < 0:
                nps_color = colors.HexColor('#EF4444')  # Red
            elif nps_data['nps_score'] < 30:
                nps_color = colors.HexColor('#F59E0B')  # Orange
            elif nps_data['nps_score'] < 50:
                nps_color = colors.HexColor('#3B82F6')  # Blue
            
            nps_table_data = [
                ['NPS Score', 'Grade', 'Promoters', 'Passives', 'Detractors'],
                [str(nps_data['nps_score']), nps_data['grade'], 
                 f"{nps_data['promoter_pct']}%", f"{nps_data['passive_pct']}%", f"{nps_data['detractor_pct']}%"],
            ]
            
            nps_table = Table(nps_table_data, colWidths=[1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch])
            nps_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
                ('TEXTCOLOR', (0, 1), (0, 1), nps_color),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ]))
            story.append(nps_table)
            
            # NPS Target comparison
            target_text = f"Target: {nps_data['target']} | Status: {nps_data['vs_target']}"
            target_color = colors.HexColor('#22C55E') if nps_data['meets_target'] else colors.HexColor('#EF4444')
            story.append(Spacer(1, 8))
            story.append(Paragraph(f"<b>{target_text}</b>", 
                ParagraphStyle('NPSTarget', parent=normal_style, fontSize=10, textColor=target_color)))
            
            # NPS explanation
            nps_explanation = """
            <i>NPS is calculated as: % Promoters - % Detractors. 
            Scores range from -100 to +100. Scores above 50 are considered excellent.</i>
            """
            story.append(Spacer(1, 8))
            story.append(Paragraph(nps_explanation, ParagraphStyle('NPSNote', parent=normal_style, fontSize=9, textColor=colors.gray)))
            story.append(Spacer(1, 20))
        
        # CSAT Score Section
        if sections.get('csatScore', True):
            story.append(Paragraph("Customer Satisfaction (CSAT)", heading_style))
            
            csat_data = self.calculate_csat(stats)
            
            # CSAT score color based on threshold
            csat_color = colors.HexColor('#22C55E') if csat_data['meets_threshold'] else colors.HexColor('#EF4444')
            
            csat_table_data = [
                ['CSAT Score', 'Status', 'Threshold', 'Satisfied', 'Unsatisfied'],
                [f"{csat_data['csat_score']}%", csat_data['status'], 
                 f"{csat_data['threshold']}%", str(csat_data['satisfied_count']), str(csat_data['unsatisfied_count'])],
            ]
            
            csat_table = Table(csat_table_data, colWidths=[1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch])
            csat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
                ('TEXTCOLOR', (0, 1), (0, 1), csat_color),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ]))
            story.append(csat_table)
            
            csat_explanation = f"""
            <i>CSAT measures customer satisfaction as % of positive feedback.
            Current threshold is set to {csat_data['threshold']}%.</i>
            """
            story.append(Spacer(1, 8))
            story.append(Paragraph(csat_explanation, ParagraphStyle('CSATNote', parent=normal_style, fontSize=9, textColor=colors.gray)))
            story.append(Spacer(1, 20))
        
        # Monthly NPS Trend Section
        if sections.get('monthlyNpsTrend', True):
            story.append(Paragraph("Monthly NPS Trend", heading_style))
            
            monthly_nps = self.get_monthly_nps(feedbacks)
            
            if len(monthly_nps) >= 2:
                # Create monthly NPS chart
                monthly_chart = self.create_monthly_nps_chart(monthly_nps)
                monthly_image = Image(io.BytesIO(monthly_chart), width=6*inch, height=3*inch)
                story.append(monthly_image)
                story.append(Spacer(1, 12))
                
                # Monthly NPS table
                monthly_table_data = [['Month', 'NPS Score', 'Total', 'Status']]
                for month in monthly_nps[-6:]:  # Last 6 months
                    status = '✓ Met Target' if month['meets_target'] else '✗ Below Target'
                    monthly_table_data.append([
                        month['month_display'],
                        str(month['nps_score']),
                        str(month['total']),
                        status
                    ])
                
                monthly_table = Table(monthly_table_data, colWidths=[1.5*inch, 1.2*inch, 1*inch, 1.5*inch])
                monthly_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
                ]))
                story.append(monthly_table)
            else:
                story.append(Paragraph("Insufficient data for monthly NPS trend (need at least 2 months)", normal_style))
            story.append(Spacer(1, 20))
        
        # Complaint Categories Section
        if sections.get('complaintCategories', True):
            story.append(Paragraph("Top Complaint Categories", heading_style))
            
            complaint_categories = self.get_complaint_categories(feedbacks, limit=8)
            
            if complaint_categories:
                complaint_table_data = [['Category', 'Count', 'Percentage']]
                
                for cat in complaint_categories:
                    complaint_table_data.append([
                        cat['category'],
                        str(cat['count']),
                        f"{cat['percentage']}%"
                    ])
                
                complaint_table = Table(complaint_table_data, colWidths=[2*inch, 1*inch, 1.2*inch])
                complaint_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
                ]))
                story.append(complaint_table)
                story.append(Spacer(1, 8))
                story.append(Paragraph("<i>Categories extracted from negative feedback using keyword detection</i>", 
                    ParagraphStyle('ComplaintNote', parent=normal_style, fontSize=9, textColor=colors.gray)))
            else:
                story.append(Paragraph("No complaint categories found in the selected period.", normal_style))
            story.append(Spacer(1, 20))
        
        # Top Routes Section
        if sections.get('topRoutes', True):
            story.append(Paragraph("Top Performing Routes", heading_style))
            
            top_routes = self.get_top_routes(feedbacks, limit=5)
            
            if top_routes:
                routes_table_data = [['Route', 'Total', 'Positive Rate', 'Wilson Score', 'Confidence']]
                
                for route in top_routes:
                    confidence = 'High' if route['total'] >= 50 else 'Medium' if route['total'] >= 20 else 'Low'
                    routes_table_data.append([
                        route['route'],
                        str(route['total']),
                        f"{route['positive_rate']}%",
                        f"{route['wilson_score']}%",
                        confidence
                    ])
                
                routes_table = Table(routes_table_data, colWidths=[1.8*inch, 0.8*inch, 1.2*inch, 1.2*inch, 1.2*inch])
                routes_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
                ]))
                story.append(routes_table)
                story.append(Spacer(1, 8))
                story.append(Paragraph("<i>Routes ranked by Wilson Score (statistical confidence of positive rate)</i>", 
                    ParagraphStyle('RouteNote', parent=normal_style, fontSize=9, textColor=colors.gray)))
            else:
                story.append(Paragraph("Insufficient route data available (minimum 5 feedbacks per route required).", normal_style))
            story.append(Spacer(1, 20))
        
        # Sentiment Distribution Chart
        if sections.get('sentimentChart', True):
            story.append(Paragraph("Sentiment Distribution", heading_style))
            
            chart_bytes = self.create_sentiment_chart(stats)
            chart_image = Image(io.BytesIO(chart_bytes), width=4*inch, height=3*inch)
            story.append(chart_image)
            story.append(Spacer(1, 20))
        
        # Trend Analysis Chart
        if sections.get('trendChart', True) and len(feedbacks) > 0:
            story.append(Paragraph("Sentiment Trends Over Time", heading_style))
            
            trend_data = self.get_trend_data(feedbacks)
            if len(trend_data['dates']) > 1:
                trend_chart_bytes = self.create_trend_chart(trend_data)
                trend_image = Image(io.BytesIO(trend_chart_bytes), width=6*inch, height=3*inch)
                story.append(trend_image)
            else:
                story.append(Paragraph("Insufficient data for trend analysis (need at least 2 dates)", normal_style))
            story.append(Spacer(1, 20))
        
        # Statistics Table
        if sections.get('statsTable', True):
            story.append(Paragraph("Detailed Statistics", heading_style))
            
            total = stats['total'] or 1
            stats_data = [
                ['Category', 'Count', 'Percentage'],
                ['Arabic Feedback', str(stats['arabic_count']), 
                 f"{round(stats['arabic_count']/total*100, 1)}%"],
                ['English Feedback', str(stats['english_count']), 
                 f"{round(stats['english_count']/total*100, 1)}%"],
                ['Mixed Feedback', str(stats.get('mixed_count', 0)), 
                 f"{round(stats.get('mixed_count', 0)/total*100, 1)}%"],
            ]
            
            stats_table = Table(stats_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 20))
        
        # Negative Feedback Samples
        if sections.get('negativeSamples', True):
            story.append(Paragraph("Top Negative Feedback Samples", heading_style))
            
            negative_samples = self.get_top_negative_feedback(feedbacks, limit=10)
            
            if negative_samples:
                # Create Arabic-compatible style
                arabic_style = ParagraphStyle(
                    'ArabicNormal',
                    parent=normal_style,
                    fontName=self.arabic_font,
                    fontSize=11,
                    spaceAfter=6,
                    wordWrap='RTL' if self.arabic_font == 'Arabic' else 'LTR'
                )
                
                for i, sample in enumerate(negative_samples, 1):
                    # Reshape Arabic text if present
                    display_text = reshape_arabic_text(sample['text'])
                    sample_text = f"<b>{i}.</b> [{sample['date']}] {display_text}"
                    story.append(Paragraph(sample_text, arabic_style))
                    story.append(Paragraph(
                        f"<i>Flight: {sample['flight']} | Customer: {sample['customer']} | Confidence: {sample['confidence']:.2f}</i>",
                        ParagraphStyle('Small', parent=normal_style, fontSize=9, textColor=colors.gray)
                    ))
                    story.append(Spacer(1, 8))
            else:
                story.append(Paragraph("No negative feedback found in the selected period.", normal_style))
        
        # Build PDF
        doc.build(story)
        
        file_size = os.path.getsize(filepath)
        return filepath, file_size
    
    def generate_excel_report(
        self,
        title: str,
        feedbacks: List[Feedback],
        stats: Dict[str, Any],
        date_from: Optional[datetime],
        date_to: Optional[datetime]
    ) -> Tuple[str, int]:
        """
        Generate an Excel report with all feedback data
        Returns: (file_path, file_size)
        """
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        
        # Summary header
        ws_summary['A1'] = title
        ws_summary['A1'].font = Font(bold=True, size=16, color="003366")
        ws_summary.merge_cells('A1:D1')
        
        date_range = ""
        if date_from and date_to:
            date_range = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        ws_summary['A2'] = f"Report Period: {date_range or 'All Time'}"
        ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Statistics table
        summary_start_row = 5
        summary_headers = ['Metric', 'Value', 'Percentage']
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=summary_start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        summary_data = [
            ('Total Feedback', stats['total'], '100%'),
            ('Positive Feedback', stats['positive'], f"{stats['positive_pct']}%"),
            ('Negative Feedback', stats['negative'], f"{stats['negative_pct']}%"),
            ('Neutral Feedback', stats['neutral'], f"{stats['neutral_pct']}%"),
            ('Arabic Feedback', stats['arabic_count'], f"{round(stats['arabic_count']/max(stats['total'],1)*100, 1)}%"),
            ('English Feedback', stats['english_count'], f"{round(stats['english_count']/max(stats['total'],1)*100, 1)}%"),
            ('Mixed Feedback', stats.get('mixed_count', 0), f"{round(stats.get('mixed_count', 0)/max(stats['total'],1)*100, 1)}%"),
            ('Average Confidence', f"{stats['avg_confidence']:.2f}", '-'),
        ]
        
        for row_idx, (metric, value, pct) in enumerate(summary_data, summary_start_row + 1):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)
            ws_summary.cell(row=row_idx, column=3, value=pct)
        
        # NPS Section
        nps_start_row = summary_start_row + len(summary_data) + 3
        nps_data = self.calculate_nps(stats)
        
        ws_summary.cell(row=nps_start_row, column=1, value="Net Promoter Score (NPS)")
        ws_summary.cell(row=nps_start_row, column=1).font = Font(bold=True, size=12, color="003366")
        
        nps_headers = ['Metric', 'Value']
        for col, header in enumerate(nps_headers, 1):
            cell = ws_summary.cell(row=nps_start_row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        nps_rows = [
            ('NPS Score', nps_data['nps_score']),
            ('Grade', nps_data['grade']),
            ('Target', nps_data['target']),
            ('Status', nps_data['vs_target']),
            ('Promoters', f"{nps_data['promoter_pct']}%"),
            ('Passives', f"{nps_data['passive_pct']}%"),
            ('Detractors', f"{nps_data['detractor_pct']}%"),
        ]
        
        for row_idx, (metric, value) in enumerate(nps_rows, nps_start_row + 2):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # CSAT Section
        csat_start_row = nps_start_row + len(nps_rows) + 3
        csat_data = self.calculate_csat(stats)
        
        ws_summary.cell(row=csat_start_row, column=1, value="Customer Satisfaction (CSAT)")
        ws_summary.cell(row=csat_start_row, column=1).font = Font(bold=True, size=12, color="003366")
        
        for col, header in enumerate(['Metric', 'Value'], 1):
            cell = ws_summary.cell(row=csat_start_row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        csat_rows = [
            ('CSAT Score', f"{csat_data['csat_score']}%"),
            ('Threshold', f"{csat_data['threshold']}%"),
            ('Status', csat_data['status']),
            ('Satisfied Count', csat_data['satisfied_count']),
            ('Unsatisfied Count', csat_data['unsatisfied_count']),
        ]
        
        for row_idx, (metric, value) in enumerate(csat_rows, csat_start_row + 2):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Analytics Settings Section
        settings_start_row = csat_start_row + len(csat_rows) + 3
        
        ws_summary.cell(row=settings_start_row, column=1, value="Analytics Settings Used")
        ws_summary.cell(row=settings_start_row, column=1).font = Font(bold=True, size=12, color="003366")
        
        settings_rows = [
            ('NPS Target', self.settings.nps_target),
            ('CSAT Threshold', f"{self.settings.csat_threshold}%"),
            ('Min Reviews per Route', self.settings.min_reviews_per_route),
        ]
        
        for row_idx, (metric, value) in enumerate(settings_rows, settings_start_row + 1):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 15
        
        # Feedback Data Sheet
        ws_data = wb.create_sheet("Feedback Data")
        
        data_headers = [
            'ID', 'Date', 'Customer Name', 'Customer Email', 'Flight Number',
            'Source', 'Language', 'Sentiment', 'Confidence', 'Status', 'Feedback Text'
        ]
        
        for col, header in enumerate(data_headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, f in enumerate(feedbacks, 2):
            ws_data.cell(row=row_idx, column=1, value=f.id)
            ws_data.cell(row=row_idx, column=2, value=f.feedback_date.strftime('%Y-%m-%d') if f.feedback_date else '')
            ws_data.cell(row=row_idx, column=3, value=f.customer_name or '')
            ws_data.cell(row=row_idx, column=4, value=f.customer_email or '')
            ws_data.cell(row=row_idx, column=5, value=f.flight_number or '')
            ws_data.cell(row=row_idx, column=6, value=f.source or '')
            ws_data.cell(row=row_idx, column=7, value=f.language or '')
            ws_data.cell(row=row_idx, column=8, value=f.sentiment or '')
            ws_data.cell(row=row_idx, column=9, value=f.sentiment_confidence or 0)
            ws_data.cell(row=row_idx, column=10, value=f.status or '')
            ws_data.cell(row=row_idx, column=11, value=f.text or '')
        
        # Adjust column widths for data sheet
        for col_idx, width in enumerate([8, 12, 20, 25, 15, 12, 10, 12, 12, 12, 50], 1):
            ws_data.column_dimensions[chr(64 + col_idx)].width = width
        
        # Trend Data Sheet
        ws_trend = wb.create_sheet("Daily Trends")
        
        trend_data = self.get_trend_data(feedbacks)
        trend_headers = ['Date', 'Total', 'Positive', 'Negative', 'Neutral']
        
        for col, header in enumerate(trend_headers, 1):
            cell = ws_trend.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, (date, total, pos, neg, neu) in enumerate(
            zip(trend_data['dates'], trend_data['total'], 
                trend_data['positive'], trend_data['negative'], trend_data['neutral']), 2):
            ws_trend.cell(row=row_idx, column=1, value=date)
            ws_trend.cell(row=row_idx, column=2, value=total)
            ws_trend.cell(row=row_idx, column=3, value=pos)
            ws_trend.cell(row=row_idx, column=4, value=neg)
            ws_trend.cell(row=row_idx, column=5, value=neu)
        
        # Monthly NPS Sheet
        ws_monthly_nps = wb.create_sheet("Monthly NPS")
        
        monthly_nps = self.get_monthly_nps(feedbacks)
        monthly_headers = ['Month', 'NPS Score', 'Total', 'Positive', 'Negative', 'Neutral', 'Meets Target']
        
        for col, header in enumerate(monthly_headers, 1):
            cell = ws_monthly_nps.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, month_data in enumerate(monthly_nps, 2):
            ws_monthly_nps.cell(row=row_idx, column=1, value=month_data['month_display'])
            ws_monthly_nps.cell(row=row_idx, column=2, value=month_data['nps_score'])
            ws_monthly_nps.cell(row=row_idx, column=3, value=month_data['total'])
            ws_monthly_nps.cell(row=row_idx, column=4, value=month_data['positive'])
            ws_monthly_nps.cell(row=row_idx, column=5, value=month_data['negative'])
            ws_monthly_nps.cell(row=row_idx, column=6, value=month_data['neutral'])
            ws_monthly_nps.cell(row=row_idx, column=7, value='Yes' if month_data['meets_target'] else 'No')
        
        # Route Performance Sheet
        ws_routes = wb.create_sheet("Route Performance")
        
        top_routes = self.get_top_routes(feedbacks, limit=20)
        route_headers = ['Route', 'Total Reviews', 'Positive Rate', 'Wilson Score', 'Confidence Level', 'Sentiment Score']
        
        for col, header in enumerate(route_headers, 1):
            cell = ws_routes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, route in enumerate(top_routes, 2):
            ws_routes.cell(row=row_idx, column=1, value=route['route'])
            ws_routes.cell(row=row_idx, column=2, value=route['total'])
            ws_routes.cell(row=row_idx, column=3, value=f"{route['positive_rate']}%")
            ws_routes.cell(row=row_idx, column=4, value=f"{route['wilson_score']}%")
            ws_routes.cell(row=row_idx, column=5, value=route.get('confidence_level', 'N/A'))
            ws_routes.cell(row=row_idx, column=6, value=f"{route.get('sentiment_score', 0)}%")
        
        # Complaint Categories Sheet
        ws_complaints = wb.create_sheet("Complaint Categories")
        
        complaint_categories = self.get_complaint_categories(feedbacks, limit=15)
        complaint_headers = ['Category', 'Count', 'Percentage of Negative']
        
        for col, header in enumerate(complaint_headers, 1):
            cell = ws_complaints.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, cat in enumerate(complaint_categories, 2):
            ws_complaints.cell(row=row_idx, column=1, value=cat['category'])
            ws_complaints.cell(row=row_idx, column=2, value=cat['count'])
            ws_complaints.cell(row=row_idx, column=3, value=f"{cat['percentage']}%")
        
        # Save workbook
        wb.save(filepath)
        
        file_size = os.path.getsize(filepath)
        return filepath, file_size
    
    def get_preview_stats(
        self,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        sentiments: Optional[List[str]] = None,
        languages: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Get preview statistics without generating a full report
        """
        # Get total count in database
        total_in_db = self.db.query(func.count(Feedback.id)).scalar()
        
        # Get filtered feedback and stats
        feedbacks, stats = self.get_filtered_feedback(date_from, date_to, sentiments, languages)
        
        return {
            'total_in_database': total_in_db,
            'matching_filters': stats['total'],
            'positive_count': stats['positive'],
            'negative_count': stats['negative'],
            'neutral_count': stats['neutral'],
            'positive_pct': stats['positive_pct'],
            'negative_pct': stats['negative_pct'],
            'neutral_pct': stats['neutral_pct'],
            'date_range': {
                'from': date_from.isoformat() if date_from else None,
                'to': date_to.isoformat() if date_to else None
            }
        }
